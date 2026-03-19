"""
Comprehensive validation suite for all 6 MMV tasks.
Runs everything that doesn't require external credentials.
"""

import sys, os, json, traceback, importlib
from datetime import datetime, timezone

BASE = os.path.dirname(os.path.abspath(__file__))

# Each repo's root contains a tools/ package. We add each root to sys.path
# but also add the tools dir directly as a fallback since modules do
# `from provenance import tag_provenance` within the same package.
for repo in ["mmv-data", "mmv-agent", "mmv-reporting", "mmv-underwriting"]:
    repo_path = os.path.join(BASE, repo)
    tools_path = os.path.join(repo_path, "tools")
    for p in [repo_path, tools_path]:
        if p not in sys.path:
            sys.path.insert(0, p)

PASS = 0
FAIL = 0
RESULTS = []

def test(name, fn):
    global PASS, FAIL
    try:
        result = fn()
        if result is True or result is None:
            PASS += 1
            RESULTS.append(("✅", name, "PASS"))
            print(f"  ✅ {name}")
        elif isinstance(result, str) and result.startswith("SKIP"):
            PASS += 1
            RESULTS.append(("⏭️", name, result))
            print(f"  ⏭️  {name}: {result}")
        else:
            FAIL += 1
            RESULTS.append(("❌", name, f"FAIL: {result}"))
            print(f"  ❌ {name}: {result}")
    except Exception as e:
        FAIL += 1
        tb = traceback.format_exc().split('\n')[-3].strip()
        RESULTS.append(("❌", name, f"ERROR: {e}"))
        print(f"  ❌ {name}: {e}")


# ════════════════════════════════════════════════════════════════
# TASK 1: Provenance Layer
# ════════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print("  TASK 1: Provenance Layer")
print("═"*60)

def test_provenance_import():
    from provenance import tag_provenance, compute_freshness, unwrap, get_provenance
    return True

def test_provenance_structure():
    from provenance import tag_provenance
    result = tag_provenance(
        [{"val": 42}],
        source="Test",
        source_url="http://test.com",
        query_params={"key": "val"},
    )
    assert "value" in result, "Missing 'value' key"
    assert "provenance" in result, "Missing 'provenance' key"
    p = result["provenance"]
    for key in ["source", "source_url", "query_params", "fetched_at", "freshness", "is_mock"]:
        assert key in p, f"Missing provenance key: {key}"
    assert p["source"] == "Test"
    assert p["is_mock"] == False
    assert p["freshness"] == "current"
    assert result["value"] == [{"val": 42}]
    return True

def test_provenance_mock_flag():
    from provenance import tag_provenance
    result = tag_provenance([], source="X", source_url="", is_mock=True)
    assert result["provenance"]["is_mock"] == True
    return True

def test_provenance_freshness_categories():
    from provenance import compute_freshness
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    assert compute_freshness(now) == "current"
    assert compute_freshness("2020-01-01T00:00:00Z") == "very_stale"
    assert compute_freshness("invalid") == "unknown"
    return True

def test_provenance_unwrap():
    from provenance import tag_provenance, unwrap, get_provenance
    tagged = tag_provenance({"x": 1}, source="S", source_url="U")
    assert unwrap(tagged) == {"x": 1}
    assert get_provenance(tagged)["source"] == "S"
    assert unwrap({"plain": "data"}) == {"plain": "data"}
    assert get_provenance({"no": "provenance"}) is None
    return True

def test_provenance_empty_data():
    from provenance import tag_provenance
    result = tag_provenance([], source="Empty", source_url="")
    assert result["value"] == []
    assert result["provenance"]["source"] == "Empty"
    return True

def _test_fetcher_import(module_name, func_names):
    """Generic test: import a fetcher module and verify functions exist."""
    mod = importlib.import_module(module_name)
    for fn in func_names:
        assert hasattr(mod, fn), f"Missing function: {fn}"
    return True

def test_fred_import():     return _test_fetcher_import("fred", ["fetch_series", "fetch_all"])
def test_usda_import():     return _test_fetcher_import("usda", ["fetch_land_values", "fetch_cash_rents", "fetch_crop_production", "fetch_crop_production_api"])
def test_eia_import():      return _test_fetcher_import("eia", ["fetch_solar_generation", "fetch_all_generation_by_fuel"])
def test_noaa_import():     return _test_fetcher_import("noaa", ["fetch_monthly_climate"])
def test_drought_import():  return _test_fetcher_import("drought", ["fetch_drought_by_state"])
def test_census_import():   return _test_fetcher_import("census", ["fetch_county_population"])
def test_ssurgo_import():   return _test_fetcher_import("ssurgo", ["fetch_soil_summary"])
def test_usdafas_import():  return _test_fetcher_import("usda_fas", ["fetch_export_sales_report"])
def test_usdafsa_import():  return _test_fetcher_import("usda_fsa", ["fetch_crp_enrollment"])

test("Provenance module imports", test_provenance_import)
test("Provenance output structure", test_provenance_structure)
test("Provenance is_mock flag", test_provenance_mock_flag)
test("Freshness categories (current/very_stale/unknown)", test_provenance_freshness_categories)
test("Unwrap and get_provenance helpers", test_provenance_unwrap)
test("Provenance wraps empty data cleanly", test_provenance_empty_data)
test("fred.py imports with provenance", test_fred_import)
test("usda.py imports with provenance + API", test_usda_import)
test("eia.py imports with provenance", test_eia_import)
test("noaa.py imports with provenance", test_noaa_import)
test("drought.py imports with provenance", test_drought_import)
test("census.py imports with provenance", test_census_import)
test("ssurgo.py imports with provenance", test_ssurgo_import)
test("usda_fas.py imports with provenance", test_usdafas_import)
test("usda_fsa.py imports with provenance", test_usdafsa_import)


# ════════════════════════════════════════════════════════════════
# TASK 2: Report Generation
# ════════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print("  TASK 2: Report Generation")
print("═"*60)

SAMPLE_UNDERWRITING = {
    "state": "TX",
    "summary": {
        "thesis": "FAVORABLE",
        "thesis_detail": "Strong signals",
        "signals": {"bullish": ["Growth +5%"], "bearish": [], "neutral": ["Moderate"]},
    },
    "sections": {
        "land_values": {
            "nominal_cagr_pct": 5.2, "real_cagr_pct": 2.1,
            "volatility_pct": 3.8,
            "period": "2018-2023",
            "start_value": 2500, "end_value": 3400,
            "yoy_changes": [{"year": 2023, "change_pct": 6.3}],
        },
        "cap_rate": {
            "latest": {"year": 2023, "cap_rate_pct": 4.18},
            "benchmark_comparison": {
                "farmland_cap_rate_pct": 4.18,
                "comparisons": {
                    "10yr_treasury": {"label": "10-Year Treasury", "spread_vs_midpoint_bps": -7, "assessment": "below"},
                },
                "summary": "Farmland at 4.18% is competitive with CRE",
            },
            "time_series": [{"year": 2023, "cap_rate_pct": 4.18}],
        },
        "risk": {
            "composite_score": 28, "risk_tier": "LOW",
            "factors": {"interest_rate_shock": {"label": "Interest Rate Shock", "score": 45, "data_gap": False}},
        },
        "crop_economics": {
            "diversification": {"hhi": 3200, "assessment": "moderately concentrated"},
            "crop_details": {
                "CORN": {"coefficient_of_variation": 8.0, "stability": "stable"},
                "COTTON": {"coefficient_of_variation": 35.0, "stability": "volatile"},
            },
            "rent_analysis": {"latest_rent_per_acre": 142, "rent_trend_pct": 9.2, "rent_stability": "rising"},
        },
        "exit_optionality": {
            "exit_tier": "STRONG", "composite_score": 78,
            "path_scores": {"solar": 85, "development": 50, "conservation": 30},
            "details": {
                "solar": {"assessment": "TX solar 600%+ growth"},
            },
        },
    },
    "data_gaps": ["Risk: Leverage data gap"],
    "data_completeness": "5/5",
}

def test_report_import():
    from markdown_report import generate_report
    return True

def test_report_generates_markdown():
    from markdown_report import generate_report
    md = generate_report(SAMPLE_UNDERWRITING)
    assert isinstance(md, str)
    assert len(md) > 500, f"Report too short: {len(md)} chars"
    return True

def test_report_has_title():
    from markdown_report import generate_report
    md = generate_report(SAMPLE_UNDERWRITING)
    assert "# Farmland Investment Analysis — TX" in md
    return True

def test_report_has_fact_labels():
    from markdown_report import generate_report
    md = generate_report(SAMPLE_UNDERWRITING)
    fact_count = md.count("`FACT`")
    assert fact_count >= 5, f"Expected >=5 FACT labels, got {fact_count}"
    return True

def test_report_has_inference_labels():
    from markdown_report import generate_report
    md = generate_report(SAMPLE_UNDERWRITING)
    assert "`INFERENCE`" in md
    return True

def test_report_has_confidence_badges():
    from markdown_report import generate_report
    md = generate_report(SAMPLE_UNDERWRITING)
    badge_count = md.count("🟢") + md.count("🟡") + md.count("🔴")
    assert badge_count >= 3, f"Expected >=3 badges, got {badge_count}"
    return True

def test_report_has_source_footnotes():
    from markdown_report import generate_report
    md = generate_report(SAMPLE_UNDERWRITING)
    assert "[^" in md, "Missing source footnotes"
    assert "## Sources" in md, "Missing Sources section"
    return True

def test_report_has_data_gaps():
    from markdown_report import generate_report
    md = generate_report(SAMPLE_UNDERWRITING)
    assert "Data Gaps" in md
    assert "Leverage" in md
    return True

def test_report_has_reproducibility():
    from markdown_report import generate_report
    md = generate_report(SAMPLE_UNDERWRITING)
    assert "Reproducibility Footer" in md
    assert "mmv-data → mmv-underwriting → mmv-reporting" in md
    return True

def test_report_all_sections():
    from markdown_report import generate_report
    md = generate_report(SAMPLE_UNDERWRITING)
    for section in ["Land Values", "Cap Rate", "Risk Scoring", "Crop Economics", "Exit Optionality"]:
        assert section in md, f"Missing section: {section}"
    return True

def test_report_empty_input():
    from markdown_report import generate_report
    md = generate_report({"state": "XX", "summary": {}, "sections": {}, "data_gaps": [], "data_completeness": "0/5"})
    assert isinstance(md, str)
    assert "XX" in md
    return True

test("Report module imports", test_report_import)
test("Generates non-trivial markdown", test_report_generates_markdown)
test("Report has correct title", test_report_has_title)
test("Report has FACT labels (>=5)", test_report_has_fact_labels)
test("Report has INFERENCE labels", test_report_has_inference_labels)
test("Report has confidence badges (>=3)", test_report_has_confidence_badges)
test("Report has source footnotes", test_report_has_source_footnotes)
test("Report has data gaps section", test_report_has_data_gaps)
test("Report has reproducibility footer", test_report_has_reproducibility)
test("Report has all 5 sections", test_report_all_sections)
test("Report handles empty input gracefully", test_report_empty_input)


# ════════════════════════════════════════════════════════════════
# TASK 3: GCS Upload
# ════════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print("  TASK 3: GCS Upload")
print("═"*60)

def test_gcs_content_hash():
    from gcs import content_hash
    import tempfile
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
        f.write('{"test": true}')
        path = f.name
    h = content_hash(path)
    assert len(h) == 64, f"SHA-256 should be 64 hex chars, got {len(h)}"
    h2 = content_hash(path)
    assert h == h2, "Same file should produce same hash"
    os.unlink(path)
    return True

def test_gcs_hash_different_content():
    from gcs import content_hash
    import tempfile
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
        f.write('{"a": 1}')
        path1 = f.name
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
        f.write('{"b": 2}')
        path2 = f.name
    h1 = content_hash(path1)
    h2 = content_hash(path2)
    assert h1 != h2, "Different content should produce different hashes"
    os.unlink(path1)
    os.unlink(path2)
    return True

def test_gcs_upload_missing_file():
    from gcs import upload_to_gcs
    result = upload_to_gcs("/nonexistent/file.json", "test-bucket")
    assert result["status"] == "error"
    return True

test("content_hash produces SHA-256", test_gcs_content_hash)
test("Different files produce different hashes", test_gcs_hash_different_content)
test("upload_to_gcs handles missing file", test_gcs_upload_missing_file)


# ════════════════════════════════════════════════════════════════
# TASK 4: Cloud SQL Tables
# ════════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print("  TASK 4: Cloud SQL Tables")
print("═"*60)

def test_ddl_file_exists():
    path = os.path.join(BASE, "mmv-infra", "sql", "create_tables.sql")
    assert os.path.exists(path), f"DDL file not found: {path}"
    with open(path) as f:
        ddl = f.read()
    assert len(ddl) > 2000, f"DDL too short: {len(ddl)} chars"
    return True

def test_ddl_has_all_tables():
    path = os.path.join(BASE, "mmv-infra", "sql", "create_tables.sql")
    with open(path) as f:
        ddl = f.read()
    tables = ["land_values", "cash_rents", "fred_series", "eia_generation",
              "noaa_climate", "drought_monitor", "usda_exports", "crp_enrollment",
              "ssurgo_soils", "census_population", "crop_production"]
    for t in tables:
        assert f"CREATE TABLE IF NOT EXISTS {t}" in ddl, f"Missing table: {t}"
    return True

def test_ddl_has_primary_keys():
    path = os.path.join(BASE, "mmv-infra", "sql", "create_tables.sql")
    with open(path) as f:
        ddl = f.read()
    pk_count = ddl.count("PRIMARY KEY")
    assert pk_count >= 11, f"Expected >=11 primary keys, got {pk_count}"
    return True

def test_loader_import():
    from load_to_postgres import TABLE_MAP, FILE_TO_TABLE
    assert len(TABLE_MAP) >= 11, f"Expected >=11 table mappings, got {len(TABLE_MAP)}"
    assert len(FILE_TO_TABLE) >= 10, f"Expected >=10 file mappings, got {len(FILE_TO_TABLE)}"
    return True

def test_loader_table_map_completeness():
    from load_to_postgres import TABLE_MAP
    for key, config in TABLE_MAP.items():
        assert "table" in config, f"Missing 'table' in {key}"
        assert "columns" in config, f"Missing 'columns' in {key}"
        assert "pk" in config, f"Missing 'pk' in {key}"
        assert len(config["columns"]) > 0, f"Empty columns for {key}"
        for pk_col in config["pk"]:
            assert pk_col in config["columns"], f"PK col '{pk_col}' not in columns for {key}"
    return True

test("DDL file exists and is non-trivial", test_ddl_file_exists)
test("DDL has all 11 tables", test_ddl_has_all_tables)
test("DDL has primary keys for all tables", test_ddl_has_primary_keys)
test("Loader module imports", test_loader_import)
test("TABLE_MAP has valid structure (PK ⊂ columns)", test_loader_table_map_completeness)


# ════════════════════════════════════════════════════════════════
# TASK 5: Crop Production API
# ════════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print("  TASK 5: Crop Production API")
print("═"*60)

def test_crop_api_function_exists():
    import usda
    assert hasattr(usda, "fetch_crop_production_api")
    assert hasattr(usda, "NASS_API_URL")
    assert usda.NASS_API_URL == "https://quickstats.nass.usda.gov/api/api_GET/"
    return True

def test_crop_api_no_key_returns_none():
    old = os.environ.pop("NASS_API_KEY", "")
    try:
        # Reimport to pick up missing key
        import usda
        importlib.reload(usda)
        result = usda.fetch_crop_production_api("TX")
    finally:
        if old:
            os.environ["NASS_API_KEY"] = old
    assert result is None, f"Should return None when no API key, got {type(result)}"
    return True

def test_crop_production_has_fallback():
    import inspect, usda
    source = inspect.getsource(usda.fetch_crop_production)
    assert "fetch_crop_production_api" in source, "Should call API function"
    assert "Falling back" in source or "qs.crops" in source, "Should have bulk TSV fallback"
    return True

test("fetch_crop_production_api() exists", test_crop_api_function_exists)
test("API returns None without key (graceful)", test_crop_api_no_key_returns_none)
test("fetch_crop_production() has API-first + fallback", test_crop_production_has_fallback)


# ════════════════════════════════════════════════════════════════
# TASK 6: LLM Reasoning Layer
# ════════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print("  TASK 6: LLM Reasoning Layer")
print("═"*60)

def test_llm_import():
    from llm_analyst import analyze_section, synthesize_report, SECTION_SCHEMA
    return True

def test_llm_deterministic_land_values():
    from llm_analyst import analyze_section
    result = analyze_section("land_values", {
        "nominal_cagr_pct": 5.2, "real_cagr_pct": 2.1,
        "volatility_pct": 3.8,
        "start_value": 2500, "end_value": 3400,
    }, force_deterministic=True)
    assert result["analysis_method"] == "deterministic"
    assert len(result["findings"]) >= 3, f"Expected >=3 findings, got {len(result['findings'])}"
    types = [f["type"] for f in result["findings"]]
    assert "FACT" in types, "Should have FACT findings"
    assert "INFERENCE" in types, "Should have INFERENCE findings"
    return True

def test_llm_deterministic_cap_rate():
    from llm_analyst import analyze_section
    result = analyze_section("cap_rate", {
        "latest": {"cap_rate_pct": 2.5},
        "benchmark_comparison": {
            "farmland_cap_rate_pct": 2.5,
            "comparisons": {"10yr_treasury": {"label": "10-Year Treasury", "spread_vs_midpoint_bps": -50, "assessment": "below"}},
        },
    }, force_deterministic=True)
    assert len(result["findings"]) >= 2
    return True

def test_llm_deterministic_risk():
    from llm_analyst import analyze_section
    result = analyze_section("risk", {
        "composite_score": 72, "risk_tier": "HIGH",
        "factors": {
            "interest_rate_shock": {"label": "Interest Rate Shock", "score": 80, "data_gap": False},
            "leverage": {"label": "Leverage", "score": 65, "data_gap": True},
        },
    }, force_deterministic=True)
    assert len(result["data_gaps"]) >= 1, "Should flag data gap for leverage"
    assert len(result["risk_flags"]) >= 1, "Should flag elevated rates"
    return True

def test_llm_deterministic_crops():
    from llm_analyst import analyze_section
    result = analyze_section("crop_economics", {
        "diversification": {"hhi": 1200, "assessment": "well-diversified"},
        "crop_details": {
            "CORN": {"coefficient_of_variation": 8.0, "stability": "stable"},
            "WHEAT": {"coefficient_of_variation": 12.0, "stability": "moderate"},
        },
    }, force_deterministic=True)
    assert len(result["findings"]) >= 3
    return True

def test_llm_deterministic_exit():
    from llm_analyst import analyze_section
    result = analyze_section("exit_optionality", {
        "exit_tier": "STRONG", "composite_score": 78,
        "path_scores": {"solar": 85, "development": 50, "conservation": 30},
        "details": {"solar": {"assessment": "Growth 600%+"}},
    }, force_deterministic=True)
    assert len(result["findings"]) >= 1
    return True

def test_llm_finding_schema():
    from llm_analyst import analyze_section
    result = analyze_section("land_values", {
        "nominal_cagr_pct": 5.2, "volatility_pct": 3.8,
        "start_value": 2500, "end_value": 3400,
    }, force_deterministic=True)
    required = {"statement", "type", "confidence", "confidence_reason", "source_citation", "data_reference"}
    for finding in result["findings"]:
        missing = required - set(finding.keys())
        assert not missing, f"Finding missing keys: {missing}"
        assert finding["type"] in ("FACT", "INFERENCE"), f"Invalid type: {finding['type']}"
        assert finding["confidence"] in ("HIGH", "MEDIUM", "LOW"), f"Invalid confidence: {finding['confidence']}"
    return True

def test_llm_synthesis():
    from llm_analyst import analyze_section, synthesize_report
    sections = {}
    for name, data in [
        ("land_values", {"nominal_cagr_pct": 5.2, "volatility_pct": 3.8, "start_value": 2500, "end_value": 3400}),
        ("risk", {"composite_score": 28, "risk_tier": "LOW", "factors": {}}),
        ("exit_optionality", {"exit_tier": "STRONG", "composite_score": 78, "path_scores": {}, "details": {}}),
    ]:
        sections[name] = analyze_section(name, data, force_deterministic=True)
    
    synthesis = synthesize_report(sections)
    assert "recommendation" in synthesis
    assert synthesis["recommendation"] in ("FAVORABLE", "MIXED", "UNFAVORABLE")
    assert "key_themes" in synthesis
    assert "synthesis_method" in synthesis
    return True

def test_llm_unknown_section():
    from llm_analyst import analyze_section
    result = analyze_section("unknown_section", {"error": "no data"}, force_deterministic=True)
    assert result["analysis_method"] == "deterministic"
    assert len(result["data_gaps"]) >= 1
    return True

def test_llm_none_values():
    """Regression: land_values with None start_value/end_value should not crash."""
    from llm_analyst import analyze_section
    result = analyze_section("land_values", {
        "nominal_cagr_pct": 3.5,
        "start_value": None,
        "end_value": None,
    }, force_deterministic=True)
    assert result["analysis_method"] == "deterministic"
    assert len(result["findings"]) >= 1
    return True

test("LLM analyst module imports", test_llm_import)
test("Land values: deterministic (>=3 findings, FACT+INFERENCE)", test_llm_deterministic_land_values)
test("Cap rate: deterministic (>=2 findings)", test_llm_deterministic_cap_rate)
test("Risk: flags data gaps and elevated scores", test_llm_deterministic_risk)
test("Crop economics: HHI + crop stability", test_llm_deterministic_crops)
test("Exit optionality: tier + components", test_llm_deterministic_exit)
test("All findings match SECTION_SCHEMA", test_llm_finding_schema)
test("Cross-section synthesis produces recommendation", test_llm_synthesis)
test("Unknown section handled gracefully", test_llm_unknown_section)
test("Regression: None values in land_values", test_llm_none_values)


# ════════════════════════════════════════════════════════════════
# TASK 7: Underwriting Analysis
# ════════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print("  TASK 7: Underwriting Analysis")
print("═"*60)

# --- Imports ---
def test_land_values_import():
    from land_values import analyze_land_values, summarize_trend, calculate_cagr
    return True

def test_cap_rate_import():
    from cap_rate import calculate_cap_rate, calculate_cap_rate_series, compare_to_benchmarks
    return True

def test_risk_scoring_import():
    from risk_scoring import score_risk, macro_risk_assessment
    return True

def test_crop_economics_import():
    from crop_economics import crop_economics_report
    return True

def test_exit_analysis_import():
    from exit_analysis import solar_exit_analysis, development_exit_analysis, exit_optionality_score
    return True

def test_underwrite_import():
    from underwrite import underwrite_deal
    return True

test("land_values module imports", test_land_values_import)
test("cap_rate module imports", test_cap_rate_import)
test("risk_scoring module imports", test_risk_scoring_import)
test("crop_economics module imports", test_crop_economics_import)
test("exit_analysis module imports", test_exit_analysis_import)
test("underwrite orchestrator imports", test_underwrite_import)

# --- Shared test fixtures ---
UW_LAND = [
    {"year": 2018, "value_per_acre": 2500, "state": "TX", "land_type": "CROPLAND"},
    {"year": 2019, "value_per_acre": 2600, "state": "TX", "land_type": "CROPLAND"},
    {"year": 2020, "value_per_acre": 2700, "state": "TX", "land_type": "CROPLAND"},
    {"year": 2021, "value_per_acre": 2900, "state": "TX", "land_type": "CROPLAND"},
    {"year": 2022, "value_per_acre": 3200, "state": "TX", "land_type": "CROPLAND"},
    {"year": 2023, "value_per_acre": 3400, "state": "TX", "land_type": "CROPLAND"},
]
UW_RENTS = [
    {"year": 2018, "rent_per_acre": 125},
    {"year": 2019, "rent_per_acre": 128},
    {"year": 2020, "rent_per_acre": 130},
    {"year": 2021, "rent_per_acre": 135},
    {"year": 2022, "rent_per_acre": 140},
    {"year": 2023, "rent_per_acre": 142},
]
UW_CROPS = [
    {"year": 2020, "crop": "CORN", "production": 800000000, "unit": "BU", "state": "TX"},
    {"year": 2021, "crop": "CORN", "production": 750000000, "unit": "BU", "state": "TX"},
    {"year": 2022, "crop": "CORN", "production": 900000000, "unit": "BU", "state": "TX"},
    {"year": 2023, "crop": "CORN", "production": 850000000, "unit": "BU", "state": "TX"},
    {"year": 2020, "crop": "COTTON", "production": 6000000, "unit": "BALES", "state": "TX"},
    {"year": 2021, "crop": "COTTON", "production": 7500000, "unit": "BALES", "state": "TX"},
    {"year": 2022, "crop": "COTTON", "production": 3500000, "unit": "BALES", "state": "TX"},
    {"year": 2023, "crop": "COTTON", "production": 5800000, "unit": "BALES", "state": "TX"},
]
UW_FRED = {
    "FEDFUNDS": [
        {"observation_date": "2025-01-01", "value": 4.33},
        {"observation_date": "2024-01-01", "value": 5.33},
    ],
    "CPIAUCSL": [
        {"observation_date": "2025-06-01", "value": 315.0},
        {"observation_date": "2024-06-01", "value": 305.0},
    ],
    "DTWEXBGS": [
        {"observation_date": "2025-06-01", "value": 128.0},
        {"observation_date": "2024-06-01", "value": 125.0},
    ],
    "FBDTLA": [
        {"observation_date": "2025-01-01", "value": 520000},
        {"observation_date": "2020-01-01", "value": 430000},
    ],
    "DGS10": [
        {"observation_date": "2025-06-01", "value": 4.25},
    ],
}
UW_EIA = [
    {"period": "2019", "state": "TX", "generation_mwh": 5000000, "fuel_type": "Solar"},
    {"period": "2020", "state": "TX", "generation_mwh": 8000000, "fuel_type": "Solar"},
    {"period": "2021", "state": "TX", "generation_mwh": 15000000, "fuel_type": "Solar"},
    {"period": "2022", "state": "TX", "generation_mwh": 25000000, "fuel_type": "Solar"},
    {"period": "2023", "state": "TX", "generation_mwh": 35000000, "fuel_type": "Solar"},
]

# --- Land Values unit tests ---
def test_lv_cagr_calculation():
    from land_values import calculate_cagr
    cagr = calculate_cagr(2500, 3400, 5)
    assert cagr is not None
    assert abs(cagr - 0.0634) < 0.001, f"Expected ~6.34% CAGR, got {cagr*100:.2f}%"
    return True

def test_lv_cagr_edge_cases():
    from land_values import calculate_cagr
    assert calculate_cagr(0, 100, 5) is None, "Zero start should return None"
    assert calculate_cagr(100, 0, 5) is None, "Zero end should return None"
    assert calculate_cagr(100, 200, 0) is None, "Zero years should return None"
    assert calculate_cagr(-100, 200, 5) is None, "Negative start should return None"
    return True

def test_lv_analyze_full():
    from land_values import analyze_land_values
    result = analyze_land_values(UW_LAND)
    assert "error" not in result
    assert result["state"] == "TX"
    assert result["period"] == "2018-2023"
    assert result["start_value"] == 2500
    assert result["end_value"] == 3400
    assert result["nominal_cagr_pct"] is not None
    assert result["nominal_cagr_pct"] > 0, "Should show positive CAGR"
    assert result["volatility_pct"] is not None
    assert result["num_observations"] == 6
    assert len(result["yoy_changes"]) == 5
    return True

def test_lv_analyze_empty():
    from land_values import analyze_land_values
    result = analyze_land_values([])
    assert "error" in result
    return True

def test_lv_analyze_single_point():
    from land_values import analyze_land_values
    result = analyze_land_values([{"year": 2023, "value_per_acre": 3400, "state": "TX"}])
    assert "error" not in result
    assert result["nominal_cagr"] is None  # Can't compute CAGR with 1 point
    assert result["num_observations"] == 1
    return True

def test_lv_real_return_with_cpi():
    from land_values import analyze_land_values
    cpi = [
        {"observation_date": "2018-06-01", "value": 251.0},
        {"observation_date": "2023-06-01", "value": 304.0},
    ]
    result = analyze_land_values(UW_LAND, cpi_data=cpi)
    assert result["real_cagr_pct"] is not None
    assert result["real_cagr_pct"] < result["nominal_cagr_pct"], "Real return should be lower than nominal"
    return True

def test_lv_summarize_trend():
    from land_values import analyze_land_values, summarize_trend
    analysis = analyze_land_values(UW_LAND)
    summary = summarize_trend(analysis)
    assert "TX" in summary
    assert "CAGR" in summary
    assert "$3,400/acre" in summary
    return True

test("CAGR: 2500→3400 over 5yr ≈ 6.3%", test_lv_cagr_calculation)
test("CAGR: edge cases (zero, negative)", test_lv_cagr_edge_cases)
test("Land values: full analysis structure", test_lv_analyze_full)
test("Land values: empty input → error dict", test_lv_analyze_empty)
test("Land values: single point → no crash", test_lv_analyze_single_point)
test("Land values: real return with CPI", test_lv_real_return_with_cpi)
test("Land values: summarize_trend output", test_lv_summarize_trend)

# --- Cap Rate unit tests ---
def test_cr_basic():
    from cap_rate import calculate_cap_rate
    result = calculate_cap_rate(3400, 142)
    assert result["cap_rate_pct"] is not None
    assert abs(result["cap_rate_pct"] - 4.18) < 0.1, f"Expected ~4.18%, got {result['cap_rate_pct']}"
    return True

def test_cr_zero_land_value():
    from cap_rate import calculate_cap_rate
    result = calculate_cap_rate(0, 142)
    assert result["cap_rate"] is None
    assert "error" in result
    return True

def test_cr_time_series():
    from cap_rate import calculate_cap_rate_series
    series = calculate_cap_rate_series(UW_LAND, UW_RENTS)
    assert len(series) == 6, f"Expected 6 years matched, got {len(series)}"
    assert series[0]["year"] == 2018
    assert series[-1]["year"] == 2023
    for entry in series:
        assert entry["cap_rate_pct"] > 0
    return True

def test_cr_time_series_no_overlap():
    from cap_rate import calculate_cap_rate_series
    future_rents = [{"year": 2030, "rent_per_acre": 200}]
    series = calculate_cap_rate_series(UW_LAND, future_rents)
    assert len(series) == 0, "No year overlap should produce empty series"
    return True

def test_cr_benchmarks():
    from cap_rate import compare_to_benchmarks
    result = compare_to_benchmarks(4.18)
    assert "comparisons" in result
    assert "summary" in result
    assert len(result["comparisons"]) == 5  # office, retail, industrial, multifamily, 10yr
    for comp in result["comparisons"].values():
        assert comp["assessment"] in ("above", "below", "in-line")
    return True

test("Cap rate: 142/3400 ≈ 4.18%", test_cr_basic)
test("Cap rate: zero land value → error", test_cr_zero_land_value)
test("Cap rate: time series join (6 years)", test_cr_time_series)
test("Cap rate: no year overlap → empty", test_cr_time_series_no_overlap)
test("Cap rate: benchmark comparison structure", test_cr_benchmarks)

# --- Risk Scoring unit tests ---
def test_risk_composite():
    from risk_scoring import score_risk
    result = score_risk(UW_FRED)
    assert "composite_score" in result
    assert 0 <= result["composite_score"] <= 100
    assert result["risk_tier"] in ("LOW", "MODERATE", "HIGH")
    assert len(result["factors"]) == 6
    return True

def test_risk_empty_fred():
    from risk_scoring import score_risk
    result = score_risk({})
    assert result["risk_tier"] in ("LOW", "MODERATE", "HIGH")
    gaps = [f["label"] for f in result["factors"].values() if f.get("data_gap")]
    assert len(gaps) >= 4, f"Expected >=4 data gaps with empty FRED, got {len(gaps)}"
    return True

def test_risk_high_rates():
    from risk_scoring import score_risk
    crisis_fred = {
        "FEDFUNDS": [
            {"observation_date": "2025-01-01", "value": 15.0},
            {"observation_date": "2024-01-01", "value": 8.0},
        ],
        "CPIAUCSL": [
            {"observation_date": "2025-06-01", "value": 360.0},
            {"observation_date": "2024-06-01", "value": 305.0},
        ],
        "DTWEXBGS": [
            {"observation_date": "2025-06-01", "value": 155.0},
            {"observation_date": "2024-06-01", "value": 125.0},
        ],
        "FBDTLA": [
            {"observation_date": "2025-01-01", "value": 700000},
            {"observation_date": "2020-01-01", "value": 430000},
        ],
        "DGS10": [{"observation_date": "2025-06-01", "value": 6.0}],
    }
    result = score_risk(crisis_fred)
    assert result["composite_score"] > 55, f"Crisis conditions should score high, got {result['composite_score']}"
    return True

def test_risk_data_gap_flags():
    from risk_scoring import score_risk
    partial_fred = {"FEDFUNDS": UW_FRED["FEDFUNDS"]}  # Only one series
    result = score_risk(partial_fred)
    gap_labels = [f["label"] for f in result["factors"].values() if f.get("data_gap")]
    assert "Trade Disruption" in gap_labels, "Trade disruption always flagged as data gap"
    return True

test("Risk scoring: composite 0-100 with tier", test_risk_composite)
test("Risk scoring: empty FRED → all data gaps", test_risk_empty_fred)
test("Risk scoring: crisis conditions → high score", test_risk_high_rates)
test("Risk scoring: data gap flag detection", test_risk_data_gap_flags)

# --- Crop Economics unit tests ---
def test_crop_econ_full():
    from crop_economics import crop_economics_report
    result = crop_economics_report(UW_CROPS, UW_RENTS)
    assert "error" not in result
    assert result["state"] == "TX"
    assert len(result["crops_analyzed"]) == 2
    assert "CORN" in result["crop_details"]
    assert "COTTON" in result["crop_details"]
    corn = result["crop_details"]["CORN"]
    assert corn["stability"] in ("stable", "moderate", "volatile")
    assert corn["coefficient_of_variation"] is not None
    return True

def test_crop_econ_hhi():
    from crop_economics import crop_economics_report
    result = crop_economics_report(UW_CROPS)
    hhi = result["diversification"]["hhi"]
    assert hhi is not None
    assert result["diversification"]["assessment"] in (
        "well-diversified", "moderately concentrated", "highly concentrated"
    )
    return True

def test_crop_econ_empty():
    from crop_economics import crop_economics_report
    result = crop_economics_report([])
    assert "error" in result
    return True

def test_crop_econ_rent_trend():
    from crop_economics import crop_economics_report
    result = crop_economics_report(UW_CROPS, UW_RENTS)
    rent = result["rent_analysis"]
    assert rent is not None
    assert rent["latest_rent_per_acre"] == 142
    assert rent["rent_stability"] in ("stable", "rising", "declining")
    return True

test("Crop economics: full report structure", test_crop_econ_full)
test("Crop economics: HHI diversification index", test_crop_econ_hhi)
test("Crop economics: empty input → error dict", test_crop_econ_empty)
test("Crop economics: rent trend analysis", test_crop_econ_rent_trend)

# --- Exit Analysis unit tests ---
def test_exit_solar():
    from exit_analysis import solar_exit_analysis
    result = solar_exit_analysis(UW_EIA, UW_LAND, UW_RENTS, "TX")
    assert result["solar_present"] == True
    assert result["generation_trend"]["total_growth_pct"] > 0
    premium = result["lease_economics"]["exit_premium_multiple"]
    assert premium is not None and premium > 1, "Solar should have premium over ag rent"
    return True

def test_exit_solar_no_data():
    from exit_analysis import solar_exit_analysis
    result = solar_exit_analysis([], state="TX")
    assert "error" in result
    return True

def test_exit_development_tiers():
    from exit_analysis import development_exit_analysis
    low = development_exit_analysis([{"year": 2023, "value_per_acre": 3000}])
    assert low["development_potential"] == "MINIMAL"
    high = development_exit_analysis([{"year": 2023, "value_per_acre": 20000}])
    assert high["development_potential"] == "HIGH"
    return True

def test_exit_composite():
    from exit_analysis import exit_optionality_score
    result = exit_optionality_score(UW_EIA, UW_LAND, UW_RENTS, "TX")
    assert result["exit_tier"] in ("STRONG", "MODERATE", "LIMITED")
    assert 0 <= result["composite_score"] <= 100
    assert "solar" in result["path_scores"]
    assert "development" in result["path_scores"]
    assert "conservation" in result["path_scores"]
    return True

test("Exit: solar analysis with premium", test_exit_solar)
test("Exit: no solar data → error dict", test_exit_solar_no_data)
test("Exit: development tier classification", test_exit_development_tiers)
test("Exit: composite optionality score", test_exit_composite)

# --- Orchestrator: underwrite_deal() ---
def test_underwrite_full_synthetic():
    from underwrite import underwrite_deal
    result = underwrite_deal("TX", UW_LAND, UW_RENTS, UW_CROPS, UW_FRED, UW_EIA)
    assert result["state"] == "TX"
    assert result["summary"]["thesis"] in ("FAVORABLE", "MIXED", "UNFAVORABLE")
    assert "signals" in result["summary"]
    for section in ["land_values", "cap_rate", "risk", "crop_economics", "exit_optionality"]:
        assert section in result["sections"], f"Missing section: {section}"
    assert isinstance(result["data_gaps"], list)
    assert "data_completeness" in result
    return True

def test_underwrite_synthesis_signals():
    from underwrite import underwrite_deal
    result = underwrite_deal("TX", UW_LAND, UW_RENTS, UW_CROPS, UW_FRED, UW_EIA)
    signals = result["summary"]["signals"]
    total = len(signals["bullish"]) + len(signals["bearish"]) + len(signals["neutral"])
    assert total >= 2, f"Expected >=2 signals, got {total}"
    return True

def test_underwrite_data_completeness():
    from underwrite import underwrite_deal
    result = underwrite_deal("TX", UW_LAND, UW_RENTS, UW_CROPS, UW_FRED, UW_EIA)
    comp = result["data_completeness"]
    assert "/5" in comp, f"Expected 'N/5' format, got {comp}"
    return True

def test_underwrite_real_data():
    real_path = "/tmp/mmv_initial_fetch/underwriting_TX.json"
    if not os.path.exists(real_path):
        return "SKIP: no underwriting_TX.json found"
    with open(real_path) as f:
        data = json.load(f)
    assert data["state"] == "TX"
    assert data["summary"]["thesis"] in ("FAVORABLE", "MIXED", "UNFAVORABLE")
    for section in ["land_values", "cap_rate", "risk", "crop_economics", "exit_optionality"]:
        assert section in data["sections"], f"Missing section: {section}"
    return True

test("Orchestrator: full underwriting with synthetic data", test_underwrite_full_synthetic)
test("Orchestrator: synthesis produces >=2 signals", test_underwrite_synthesis_signals)
test("Orchestrator: data completeness format", test_underwrite_data_completeness)
test("Orchestrator: real TX data structure (if available)", test_underwrite_real_data)


# ════════════════════════════════════════════════════════════════
# INTEGRATION: FRED live fetch with provenance
# ════════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print("  INTEGRATION: FRED live fetch with provenance")
print("═"*60)

def test_fred_live_fetch():
    from fred import fetch_series
    result = fetch_series("FEDFUNDS", start_date="2024-01-01")
    assert isinstance(result, dict), f"Expected dict, got {type(result)}"
    assert "value" in result, "Missing 'value' key"
    assert "provenance" in result, "Missing 'provenance' key"
    records = result["value"]
    assert len(records) > 0, "No records returned"
    assert records[0].get("series_id") == "FEDFUNDS"
    p = result["provenance"]
    assert p["source"].startswith("FRED")
    assert "fetched_at" in p
    assert p["freshness"] == "current"
    assert p["is_mock"] == False
    return True

test("FRED: live fetch returns provenance-wrapped data", test_fred_live_fetch)


# ════════════════════════════════════════════════════════════════
# INTEGRATION: Report + LLM from real underwriting data
# ════════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print("  INTEGRATION: Report + LLM from real underwriting data")
print("═"*60)

def test_report_from_real_data():
    real_path = "/tmp/mmv_initial_fetch/underwriting_TX.json"
    if not os.path.exists(real_path):
        return "SKIP: no underwriting_TX.json found"
    with open(real_path) as f:
        data = json.load(f)
    from markdown_report import generate_report
    md = generate_report(data)
    assert len(md) > 200
    assert "TX" in md
    with open("/tmp/mmv_report_TX_test.md", "w") as f:
        f.write(md)
    return True

def test_llm_from_real_data():
    real_path = "/tmp/mmv_initial_fetch/underwriting_TX.json"
    if not os.path.exists(real_path):
        return "SKIP: no underwriting_TX.json found"
    with open(real_path) as f:
        data = json.load(f)
    from llm_analyst import analyze_section, synthesize_report
    sections = data.get("sections", {})
    analyzed = {}
    for name, sdata in sections.items():
        analyzed[name] = analyze_section(name, sdata, force_deterministic=True)
    synthesis = synthesize_report(analyzed)
    assert synthesis["recommendation"] in ("FAVORABLE", "MIXED", "UNFAVORABLE")
    total = sum(len(a.get("findings", [])) for a in analyzed.values())
    assert total > 0, "Expected some findings"
    return True

test("Report from real underwriting_TX.json", test_report_from_real_data)
test("LLM analyst from real underwriting_TX.json", test_llm_from_real_data)


# ════════════════════════════════════════════════════════════════
# TASK 8: Commercial Property Tools
# ════════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print("  TASK 8: Commercial Property Tools")
print("═"*60)

# --- Imports ---
def test_cad_import():
    return _test_fetcher_import("cad", [
        "fetch_commercial_properties", "lookup_property", "search_by_owner",
        "COUNTY_CONFIG", "COMMERCIAL_SPTB_CODES",
    ])

def test_tx_entities_import():
    return _test_fetcher_import("tx_entities", [
        "search_entity", "get_entity_details", "check_franchise_tax_status",
    ])

def test_sec_edgar_import():
    return _test_fetcher_import("sec_edgar", [
        "search_filings", "get_company_filings", "search_tx_reits",
        "build_user_agent", "TX_COMMERCIAL_REITS",
    ])

def test_fema_import():
    return _test_fetcher_import("fema", [
        "lookup_flood_zone", "lookup_flood_zone_by_address",
        "classify_flood_risk", "FLOOD_ZONE_RISK",
    ])

def test_tamu_import():
    return _test_fetcher_import("tamu_realestate", [
        "fetch_market_reports", "list_available_reports",
        "build_report_url", "REPORT_CATALOG",
    ])

test("cad.py imports (CAD property data)", test_cad_import)
test("tx_entities.py imports (entity research)", test_tx_entities_import)
test("sec_edgar.py imports (SEC EDGAR)", test_sec_edgar_import)
test("fema.py imports (flood zone)", test_fema_import)
test("tamu_realestate.py imports (TX A&M)", test_tamu_import)

# --- CAD Parser Tests ---
def test_cad_parse_hcad_real_acct():
    from cad import _parse_hcad_real_acct
    mock_tsv = "ACCT\tYR\tOWNER\tSITE_ADDR_1\tSTATE_CLASS\tLAND_VAL\tBLDG_VAL\tTOT_MKT_VAL\tTOT_APPR_VAL\tASSESSED_VAL\tX_FEATURES_VAL\tADDR1\tADDR2\tCITY\tSTATE\tZIP\tSITE_ADDR_2\tSITE_ADDR_3\tLEGAL1\tLEGAL2\n"
    mock_tsv += "1234567890123\t2025\tBROOKFIELD ASSET MGMT\t1000 MAIN ST\tF1\t5000000\t15000000\t20000000\t19500000\t19500000\t500000\t\t\t\t\t\t\t\tLOT 1 BLK 2\tSOME ADDITION\n"
    mock_tsv += "9876543210987\t2025\tSMITH JOHN\t500 ELM ST\tA1\t200000\t150000\t350000\t340000\t340000\t0\t\t\t\t\t\t\t\tLOT 5\t\n"

    results = _parse_hcad_real_acct(mock_tsv)
    assert len(results) == 2, f"Expected 2 records, got {len(results)}"
    assert results[0]["account_number"] == "1234567890123"
    assert results[0]["owner_name"] == "BROOKFIELD ASSET MGMT"
    assert results[0]["sptb_code"] == "F1"
    assert results[0]["total_market_value"] == 20000000
    assert results[0]["land_value"] == 5000000
    assert results[0]["building_value"] == 15000000
    return True

def test_cad_classify_property_type():
    from cad import _classify_property_type
    assert _classify_property_type("F1") == "commercial"
    assert _classify_property_type("F1A") == "commercial"
    assert _classify_property_type("F2") == "industrial"
    assert _classify_property_type("B1") == "multifamily"
    assert _classify_property_type("A1") == "residential"
    assert _classify_property_type("D1") == "agricultural"
    assert _classify_property_type("C1") == "vacant_land"
    assert _classify_property_type("") == "unknown"
    return True

def test_cad_parse_building_other():
    from cad import _parse_hcad_building_other
    mock_tsv = "ACCT\tBLD_NUM\tIMPR_TP\tYR_BUILT\tGROSS_AR\tEFF_AR\tSTORIES\tCOND\tGRADE\n"
    mock_tsv += "1234567890123\t1\tOFFICE\t1998\t50000\t48000\t5\tGOOD\tA\n"
    mock_tsv += "1234567890123\t2\tPARKING\t1998\t20000\t20000\t3\tAVG\tB\n"

    buildings = _parse_hcad_building_other(mock_tsv)
    assert "1234567890123" in buildings
    assert len(buildings["1234567890123"]) == 2
    assert buildings["1234567890123"][0]["year_built"] == 1998
    assert buildings["1234567890123"][0]["gross_area_sqft"] == 50000
    assert buildings["1234567890123"][0]["num_stories"] == 5.0
    assert buildings["1234567890123"][0]["grade"] == "A"
    return True

def test_cad_safe_int():
    from cad import _safe_int, _safe_float
    assert _safe_int("1,234,567") == 1234567
    assert _safe_int("(D)") is None
    assert _safe_int(None) is None
    assert _safe_int("") is None
    assert _safe_int("3.14") == 3
    assert _safe_float("3.14") == 3.14
    assert _safe_float(None) is None
    return True

def test_cad_county_config():
    from cad import COUNTY_CONFIG
    assert "harris" in COUNTY_CONFIG
    assert COUNTY_CONFIG["harris"]["has_bulk"] == True
    assert "travis" in COUNTY_CONFIG
    assert COUNTY_CONFIG["travis"]["has_bulk"] == False
    for county, config in COUNTY_CONFIG.items():
        assert "name" in config, f"Missing 'name' in {county}"
    return True

def test_cad_unsupported_county():
    from cad import fetch_commercial_properties
    result = fetch_commercial_properties(county="nonexistent")
    assert isinstance(result, dict)
    assert "provenance" in result
    data = result.get("value", [])
    assert len(data) == 0
    return True

test("CAD: parse HCAD real_acct mock TSV", test_cad_parse_hcad_real_acct)
test("CAD: classify SPTB property types", test_cad_classify_property_type)
test("CAD: parse building_other mock TSV", test_cad_parse_building_other)
test("CAD: _safe_int and _safe_float edge cases", test_cad_safe_int)
test("CAD: county config completeness", test_cad_county_config)
test("CAD: unsupported county returns empty", test_cad_unsupported_county)

# --- Entity Research Tests ---
def test_entities_no_key_graceful():
    old = os.environ.pop("OPENCORPORATES_API_KEY", "")
    try:
        import tx_entities
        importlib.reload(tx_entities)
        result = tx_entities.search_entity("TEST LLC")
    finally:
        if old:
            os.environ["OPENCORPORATES_API_KEY"] = old
    assert isinstance(result, dict)
    assert "provenance" in result
    data = result.get("value", [])
    assert len(data) == 0, "Should return empty without API key"
    return True

def test_entities_parse_comptroller():
    from tx_entities import _parse_comptroller_results
    mock_html = """
    <table>
    <tr><td>12345678901</td><td>TEST LLC</td><td>HOUSTON TX</td><td>Active</td></tr>
    <tr><td>98765432109</td><td>ANOTHER LLC</td><td>DALLAS TX</td><td>Forfeited</td></tr>
    </table>
    """
    results = _parse_comptroller_results(mock_html, "TEST LLC")
    assert len(results) == 2
    assert results[0]["taxpayer_number"] == "12345678901"
    assert results[0]["taxpayer_name"] == "TEST LLC"
    assert results[0]["right_to_transact"] == "Active"
    assert results[1]["right_to_transact"] == "Forfeited"
    return True

def test_entities_empty_comptroller():
    from tx_entities import _parse_comptroller_results
    results = _parse_comptroller_results("No matching records found", "NONEXISTENT")
    assert len(results) == 0
    return True

test("Entities: no API key → graceful empty", test_entities_no_key_graceful)
test("Entities: parse Comptroller HTML results", test_entities_parse_comptroller)
test("Entities: empty Comptroller response", test_entities_empty_comptroller)

# --- SEC EDGAR Tests ---
def test_edgar_user_agent():
    from sec_edgar import build_user_agent, SEC_USER_AGENT
    ua = build_user_agent("TestApp", "test@example.com")
    assert "TestApp" in ua
    assert "test@example.com" in ua
    assert "(" in ua and ")" in ua, "SEC requires email in parentheses"
    assert "MMV-Data-Pipeline" in SEC_USER_AGENT
    return True

def test_edgar_reits_catalog():
    from sec_edgar import TX_COMMERCIAL_REITS
    assert len(TX_COMMERCIAL_REITS) >= 3
    for cik, info in TX_COMMERCIAL_REITS.items():
        assert "name" in info, f"Missing 'name' for CIK {cik}"
        assert "type" in info, f"Missing 'type' for CIK {cik}"
        assert len(cik) >= 7, f"CIK too short: {cik}"
    return True

test("EDGAR: User-Agent format compliance", test_edgar_user_agent)
test("EDGAR: TX REIT catalog completeness", test_edgar_reits_catalog)

# --- FEMA Flood Zone Tests ---
def test_fema_classify_risk():
    from fema import classify_flood_risk, FLOOD_ZONE_RISK
    assert classify_flood_risk("AE") == "high"
    assert classify_flood_risk("VE") == "very_high"
    assert classify_flood_risk("X") == "low"
    assert classify_flood_risk("A") == "high"
    assert classify_flood_risk("D") == "unknown"
    assert classify_flood_risk("") == "unknown"
    assert classify_flood_risk("NONEXISTENT") == "unknown"
    return True

def test_fema_zone_catalog():
    from fema import FLOOD_ZONE_RISK
    assert len(FLOOD_ZONE_RISK) >= 10
    for zone, info in FLOOD_ZONE_RISK.items():
        assert "risk" in info, f"Missing 'risk' for zone {zone}"
        assert "description" in info, f"Missing 'description' for zone {zone}"
        assert info["risk"] in ("very_high", "high", "moderate", "low", "unknown", "unmapped")
    return True

test("FEMA: flood risk classification", test_fema_classify_risk)
test("FEMA: zone catalog completeness", test_fema_zone_catalog)

# --- TX A&M Real Estate Center Tests ---
def test_tamu_report_catalog():
    from tamu_realestate import REPORT_CATALOG, RECENTER_BASE_URL
    assert len(REPORT_CATALOG) >= 5
    for key, info in REPORT_CATALOG.items():
        assert "url" in info, f"Missing 'url' for {key}"
        assert "description" in info, f"Missing 'description' for {key}"
        assert info["url"].startswith("https://"), f"Invalid URL for {key}"
    return True

def test_tamu_build_url():
    from tamu_realestate import build_report_url
    url = build_report_url("Houston", "office")
    assert "houston" in url
    assert "office" in url
    assert url.startswith("https://")

    url2 = build_report_url("San Antonio", "retail")
    assert "san-antonio" in url2
    assert "retail" in url2
    return True

def test_tamu_list_reports():
    from tamu_realestate import list_available_reports
    result = list_available_reports()
    assert isinstance(result, dict)
    assert "provenance" in result
    data = result.get("value", [])
    assert len(data) >= 5
    for r in data:
        assert "report_key" in r
        assert "url" in r
    return True

def test_tamu_table_extractor():
    from tamu_realestate import _extract_tables
    html = """
    <table>
    <tr><th>Metro</th><th>Vacancy</th><th>Rent</th></tr>
    <tr><td>Houston</td><td>18.5%</td><td>$32.50</td></tr>
    <tr><td>Dallas</td><td>15.2%</td><td>$35.00</td></tr>
    </table>
    """
    tables = _extract_tables(html)
    assert len(tables) == 1
    assert tables[0]["headers"] == ["Metro", "Vacancy", "Rent"]
    assert len(tables[0]["data"]) == 2
    assert tables[0]["data"][0][0] == "Houston"
    return True

def test_tamu_metric_parser():
    from tamu_realestate import _parse_market_report
    mock_html = """
    <html><head><title>Houston Office Market Q4 2025</title></head>
    <body><article>
    <p>The Houston office vacancy rate was 19.2 percent in Q4 2025,
    with an average lease rate of $31.50 per square foot.
    Net absorption totaled 1,250,000 square feet.
    The capitalization rate averaged 7.5 percent.</p>
    </article></body></html>
    """
    result = _parse_market_report(mock_html, "Houston", "office")
    assert result["title"] == "Houston Office Market Q4 2025"
    metrics = result.get("metrics", {})
    assert metrics.get("vacancy_rate_pct") == 19.2, f"Expected 19.2, got {metrics.get('vacancy_rate_pct')}"
    assert metrics.get("avg_lease_rate_psf") == 31.50
    assert metrics.get("absorption_sqft") == 1250000
    assert metrics.get("cap_rate_pct") == 7.5
    return True

test("TAMU: report catalog completeness", test_tamu_report_catalog)
test("TAMU: URL builder for metros", test_tamu_build_url)
test("TAMU: list_available_reports structure", test_tamu_list_reports)
test("TAMU: HTML table extractor", test_tamu_table_extractor)
test("TAMU: metric parser (vacancy, rent, absorption, cap rate)", test_tamu_metric_parser)



# ════════════════════════════════════════════════════════════════
# TASK 8: Excel Model Integration
# ════════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print("  TASK 8: Excel Model Parser")
print("═"*60)

import tempfile

def _create_test_workbook(path=None):
    """Create a standard test workbook for parser/mapper/executor tests."""
    import openpyxl as _xl
    wb = _xl.Workbook()
    ws = wb.active
    ws.title = "Deal Model"
    ws["A1"] = "Test Deal Model"
    ws["B4"] = "Purchase Price"
    ws["C4"] = 1200000
    ws["C4"].number_format = "$#,##0"
    ws["B5"] = "Total Acres"
    ws["C5"] = 500
    ws["B6"] = "Cash Rent ($/acre)"
    ws["C6"] = 185
    ws["B7"] = "Appreciation Rate"
    ws["C7"] = 0.03
    ws["C7"].number_format = "0.00%"
    ws["B8"] = "Discount Rate"
    ws["C8"] = 0.08
    ws["C8"].number_format = "0.00%"
    ws["B11"] = "Price per Acre"
    ws["C11"] = "=C4/C5"
    ws["B12"] = "Gross Income"
    ws["C12"] = "=C6*C5"
    ws["B13"] = "Cap Rate"
    ws["C13"] = "=C12/C4"
    ws["C13"].number_format = "0.00%"
    ws["B14"] = "NOI"
    ws["C14"] = "=C12*0.85"
    if path is None:
        path = os.path.join(tempfile.gettempdir(), "mmv_test_model.xlsx")
    wb.save(path)
    wb.close()
    return path

def test_parser_import():
    from excel_parser import parse_excel_model
    return True

def test_parser_basic():
    from excel_parser import parse_excel_model
    path = _create_test_workbook()
    result = parse_excel_model(path)
    assert "error" not in result, f"Parser error: {result.get('error')}"
    assert result["summary"]["input_count"] == 5, f"Expected 5 inputs, got {result['summary']['input_count']}"
    assert result["summary"]["formula_count"] == 4, f"Expected 4 formulas, got {result['summary']['formula_count']}"
    return True

def test_parser_classifies_inputs():
    from excel_parser import parse_excel_model
    path = _create_test_workbook()
    result = parse_excel_model(path)
    input_addrs = [c["address"] for c in result["input_cells"]]
    for expected in ["C4", "C5", "C6", "C7", "C8"]:
        assert expected in input_addrs, f"Missing input: {expected}"
    return True

def test_parser_classifies_formulas():
    from excel_parser import parse_excel_model
    path = _create_test_workbook()
    result = parse_excel_model(path)
    formula_addrs = [c["address"] for c in result["formula_cells"]]
    for expected in ["C11", "C12", "C13", "C14"]:
        assert expected in formula_addrs, f"Missing formula: {expected}"
    return True

def test_parser_extracts_labels():
    from excel_parser import parse_excel_model
    path = _create_test_workbook()
    result = parse_excel_model(path)
    labels = {c["address"]: c.get("label") for c in result["input_cells"]}
    assert labels.get("C4") == "Purchase Price", f"C4 label: {labels.get('C4')}"
    assert labels.get("C6") == "Cash Rent ($/acre)", f"C6 label: {labels.get('C6')}"
    return True

def test_parser_missing_file():
    from excel_parser import parse_excel_model
    result = parse_excel_model("/nonexistent/model.xlsx")
    assert "error" in result
    return True

def test_parser_wrong_extension():
    from excel_parser import parse_excel_model
    result = parse_excel_model("/tmp/test.csv")
    assert "error" in result
    assert "Unsupported" in result["error"]
    return True

def test_parser_multisheet():
    import openpyxl as _xl
    wb = _xl.Workbook()
    ws1 = wb.active
    ws1.title = "Assumptions"
    ws1["A1"] = "Rate"
    ws1["B1"] = 0.05
    ws2 = wb.create_sheet("Outputs")
    ws2["A1"] = "Result"
    ws2["B1"] = "=Assumptions!B1*2"
    path = os.path.join(tempfile.gettempdir(), "mmv_multi.xlsx")
    wb.save(path)
    wb.close()
    from excel_parser import parse_excel_model
    result = parse_excel_model(path)
    assert result["summary"]["total_sheets"] == 2, f"Expected 2 sheets, got {result['summary']['total_sheets']}"
    return True

test("Excel parser module imports", test_parser_import)
test("Parser: basic .xlsx with values and formulas", test_parser_basic)
test("Parser: correctly classifies input cells", test_parser_classifies_inputs)
test("Parser: correctly classifies formula cells", test_parser_classifies_formulas)
test("Parser: extracts labels from adjacent text", test_parser_extracts_labels)
test("Parser: handles missing file gracefully", test_parser_missing_file)
test("Parser: rejects non-.xlsx files", test_parser_wrong_extension)
test("Parser: multi-sheet workbook", test_parser_multisheet)


print("\n" + "═"*60)
print("  TASK 9: Excel Model Mapper")
print("═"*60)

def test_mapper_import():
    from model_mapper import map_excel_model, CONCEPT_VOCABULARY
    assert len(CONCEPT_VOCABULARY) >= 25, f"Expected >= 25 concepts, got {len(CONCEPT_VOCABULARY)}"
    return True

def test_mapper_purchase_price():
    from model_mapper import map_excel_model
    skeleton = {"input_cells": [{"address": "C4", "value": 1000000, "label": "Purchase Price", "sheet": "S"}], "formula_cells": []}
    result = map_excel_model(skeleton, force_deterministic=True)
    assert result["mapped_inputs"]["C4"]["concept"] == "purchase_price"
    return True

def test_mapper_cash_rent():
    from model_mapper import map_excel_model
    skeleton = {"input_cells": [{"address": "C6", "value": 185, "label": "Cash Rent ($/acre)", "sheet": "S"}], "formula_cells": []}
    result = map_excel_model(skeleton, force_deterministic=True)
    assert result["mapped_inputs"]["C6"]["concept"] == "cash_rent_per_acre", \
        f"Got {result['mapped_inputs']['C6']['concept']} instead of cash_rent_per_acre"
    return True

def test_mapper_irr_formula():
    from model_mapper import map_excel_model
    skeleton = {"input_cells": [], "formula_cells": [{"address": "D10", "formula": "=XIRR(C2:C10,B2:B10)", "label": None, "sheet": "S"}]}
    result = map_excel_model(skeleton, force_deterministic=True)
    assert result["mapped_outputs"]["D10"]["concept"] == "irr"
    return True

def test_mapper_model_type():
    from model_mapper import map_excel_model
    skeleton = {
        "input_cells": [{"address": "C4", "value": 100, "label": "Purchase Price", "sheet": "S"}],
        "formula_cells": [{"address": "C10", "formula": "=C5/C4", "label": "Cap Rate", "sheet": "S"}],
    }
    result = map_excel_model(skeleton, force_deterministic=True)
    assert "direct_capitalization" in result["model_type"]
    return True

def test_mapper_mmv_sources():
    from model_mapper import map_excel_model
    skeleton = {"input_cells": [
        {"address": "C4", "value": 185, "label": "Cash Rent", "sheet": "S"},
        {"address": "C5", "value": 0.08, "label": "Discount Rate", "sheet": "S"},
    ], "formula_cells": []}
    result = map_excel_model(skeleton, force_deterministic=True)
    tool_names = [s["tool"] for s in result["mmv_data_sources"]]
    assert "fetch_cash_rents" in tool_names, f"Missing fetch_cash_rents in {tool_names}"
    assert "fetch_interest_rates" in tool_names, f"Missing fetch_interest_rates in {tool_names}"
    return True

def test_mapper_unknown_label():
    from model_mapper import map_excel_model
    skeleton = {"input_cells": [{"address": "C4", "value": 42, "label": "ZZZ Unknown Field", "sheet": "S"}], "formula_cells": []}
    result = map_excel_model(skeleton, force_deterministic=True)
    assert len(result["unmapped_cells"]) == 1
    assert result["total_unmapped"] == 1
    return True

def test_mapper_empty_skeleton():
    from model_mapper import map_excel_model
    result = map_excel_model({"input_cells": [], "formula_cells": []})
    assert "error" in result
    return True

test("Mapper module imports + vocabulary", test_mapper_import)
test("Mapper: 'Purchase Price' → purchase_price", test_mapper_purchase_price)
test("Mapper: 'Cash Rent ($/acre)' → cash_rent_per_acre", test_mapper_cash_rent)
test("Mapper: =XIRR() formula → irr", test_mapper_irr_formula)
test("Mapper: detects direct_capitalization model type", test_mapper_model_type)
test("Mapper: maps concepts to MMV data sources", test_mapper_mmv_sources)
test("Mapper: unknown label → unmapped_cells", test_mapper_unknown_label)
test("Mapper: empty skeleton → error", test_mapper_empty_skeleton)


print("\n" + "═"*60)
print("  TASK 10: Excel Model Executor")
print("═"*60)

def test_executor_import():
    from model_executor import populate_excel_model, read_excel_outputs
    return True

def test_executor_populates():
    from model_executor import populate_excel_model
    path = _create_test_workbook()
    schema = {"mapped_inputs": {
        "C4": {"concept": "purchase_price", "value": 1200000, "sheet": "Deal Model"},
        "C6": {"concept": "cash_rent_per_acre", "value": 185, "sheet": "Deal Model"},
    }, "mapped_outputs": {}}
    data = {"purchase_price": 1500000, "cash_rent_per_acre": 195}
    result = populate_excel_model(path, schema, data)
    assert "error" not in result, f"Executor error: {result.get('error')}"
    assert result["summary"]["written"] == 2
    assert os.path.exists(result["output_file"])
    return True

def test_executor_preserves_formulas():
    import openpyxl as _xl
    from model_executor import populate_excel_model
    path = _create_test_workbook()
    schema = {"mapped_inputs": {"C4": {"concept": "purchase_price", "value": 1200000, "sheet": "Deal Model"}}, "mapped_outputs": {}}
    data = {"purchase_price": 1500000}
    result = populate_excel_model(path, schema, data)
    wb = _xl.load_workbook(result["output_file"])
    ws = wb["Deal Model"]
    assert ws["C4"].value == 1500000, f"C4 should be 1500000, got {ws['C4'].value}"
    assert ws["C11"].value == "=C4/C5", f"C11 formula should be preserved, got {ws['C11'].value}"
    wb.close()
    return True

def test_executor_skips_missing_data():
    from model_executor import populate_excel_model
    path = _create_test_workbook()
    schema = {"mapped_inputs": {
        "C4": {"concept": "purchase_price", "value": 1200000, "sheet": "Deal Model"},
        "C5": {"concept": "total_acres", "value": 500, "sheet": "Deal Model"},
    }, "mapped_outputs": {}}
    data = {"purchase_price": 1500000}  # total_acres not provided
    result = populate_excel_model(path, schema, data)
    assert result["summary"]["written"] == 1
    assert result["summary"]["skipped"] == 1
    assert result["cells_skipped"][0]["concept"] == "total_acres"
    return True

def test_executor_missing_file():
    from model_executor import populate_excel_model
    result = populate_excel_model("/nonexistent.xlsx", {"mapped_inputs": {}}, {})
    assert "error" in result
    return True

def test_executor_no_inputs():
    from model_executor import populate_excel_model
    path = _create_test_workbook()
    result = populate_excel_model(path, {"mapped_inputs": {}}, {})
    assert "error" in result
    return True

def test_executor_old_values_tracked():
    from model_executor import populate_excel_model
    path = _create_test_workbook()
    schema = {"mapped_inputs": {"C4": {"concept": "purchase_price", "value": 1200000, "sheet": "Deal Model"}}, "mapped_outputs": {}}
    data = {"purchase_price": 1500000}
    result = populate_excel_model(path, schema, data)
    cell = result["cells_written"][0]
    assert cell["old_value"] == 1200000, f"Old value should be tracked: {cell}"
    assert cell["new_value"] == 1500000
    return True

test("Executor module imports", test_executor_import)
test("Executor: populates input cells", test_executor_populates)
test("Executor: preserves formula cells", test_executor_preserves_formulas)
test("Executor: skips concepts with missing data", test_executor_skips_missing_data)
test("Executor: handles missing file", test_executor_missing_file)
test("Executor: rejects empty mapped_inputs", test_executor_no_inputs)
test("Executor: tracks old → new values", test_executor_old_values_tracked)


# ════════════════════════════════════════════════════════════════
# SUMMARY
# ════════════════════════════════════════════════════════════════
print("\n" + "═"*60)
print(f"  RESULTS: {PASS} passed, {FAIL} failed ({PASS+FAIL} total)")
print("═"*60)

if FAIL > 0:
    print("\n  FAILURES:")
    for icon, name, detail in RESULTS:
        if icon == "❌":
            print(f"    {icon} {name}: {detail}")

print()
sys.exit(1 if FAIL > 0 else 0)
